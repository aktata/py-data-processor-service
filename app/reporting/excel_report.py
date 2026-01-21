from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

RISK_COLORS = {
    "low": "C6EFCE",
    "medium": "FFEB9C",
    "high": "FFC7CE",
}


def export_excel_report(
    metrics_df: pd.DataFrame,
    ranking_df: pd.DataFrame,
    drilldown_df: pd.DataFrame | None,
    output_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metrics_df.to_excel(writer, sheet_name="指标表", index=False)
        ranking_df.to_excel(writer, sheet_name="排名表", index=False)
        if drilldown_df is not None:
            drilldown_df.to_excel(writer, sheet_name="明细下钻", index=False)

        worksheet = writer.sheets["指标表"]
        risk_col_idx = metrics_df.columns.get_loc("risk_level") + 1
        for row_idx in range(2, len(metrics_df) + 2):
            cell = worksheet.cell(row=row_idx, column=risk_col_idx)
            color = RISK_COLORS.get(cell.value)
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    return output_path
